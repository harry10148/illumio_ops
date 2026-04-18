"""openpyxl-based xlsx export for illumio_ops reports.

One sheet per analysis module + a Summary sheet. Header row frozen,
alternate-row banding for readability, red fill on 'blocked' / 'deny'
rows. chart_spec (if present) rendered as matplotlib PNG and embedded.
"""
from __future__ import annotations

import io
from loguru import logger
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.report.exporters.chart_renderer import render_matplotlib_png

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="375379")
_ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")
_ALERT_TOKENS = ("blocked", "deny", "violat", "critical", "red_flag")

def _write_module_sheet(wb: Workbook, name: str, module_data: dict[str, Any]) -> None:
    # openpyxl sheet names capped at 31 chars and cannot contain :\/?*[]
    safe_name = "".join(c for c in name if c not in r"\/:?*[]")[:31] or "Sheet"
    ws = wb.create_sheet(title=safe_name)

    row = 1
    summary = module_data.get("summary")
    if summary:
        ws.cell(row=row, column=1, value=str(summary)).font = Font(italic=True)
        row += 2

    table = module_data.get("table") or []
    if table:
        # Write header
        headers = list(table[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=str(header))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        header_row = row
        row += 1

        # Write data rows
        for data_row in table:
            row_vals = [data_row.get(h, "") for h in headers]
            # Highlight "blocked / deny" rows
            row_text = " ".join(str(v).lower() for v in row_vals)
            is_alert = any(tok in row_text for tok in _ALERT_TOKENS)
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if is_alert:
                    cell.fill = _ALERT_FILL
            row += 1

        ws.freeze_panes = f"A{header_row + 1}"

        # Auto-size columns (rough heuristic)
        for col_idx, header in enumerate(headers, 1):
            max_len = max(
                len(str(header)),
                max((len(str(data_row.get(header, ""))) for data_row in table), default=0),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    chart_spec = module_data.get("chart_spec")
    if chart_spec:
        try:
            png = render_matplotlib_png(chart_spec)
            img = XLImage(io.BytesIO(png))
            img.anchor = f"A{row + 2}"
            ws.add_image(img)
        except Exception as exc:
            logger.warning("Failed to render chart for {}: {}", safe_name, exc)

def export_xlsx(report_result: dict[str, Any], output_path: str) -> None:
    """Export a ReportResult-shaped dict to an .xlsx file."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    meta = report_result.get("metadata", {})
    summary_ws["A1"] = meta.get("title", "Report")
    summary_ws["A1"].font = Font(size=18, bold=True)
    summary_ws["A2"] = f"Generated: {meta.get('generated_at', '')}"
    if meta.get("start_date"):
        summary_ws["A3"] = f"Period: {meta.get('start_date')} \u2192 {meta.get('end_date', '')}"
    summary_ws["A4"] = f"Records: {report_result.get('record_count', 0)}"
    summary_ws.freeze_panes = "A2"

    for mod_name, mod_data in (report_result.get("module_results") or {}).items():
        _write_module_sheet(wb, mod_name, mod_data)

    wb.save(output_path)
    logger.info("xlsx report written to {}", output_path)
