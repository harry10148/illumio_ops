"""
src/report/exporters/excel_exporter.py
Exports a ReportResult to a multi-sheet Excel workbook using openpyxl.

Sheet order matches the plan:
  1  Executive Summary   ← mod12
  2  Traffic Overview    ← mod01
  3  Policy Decisions    ← mod02
  4  Uncovered Flows     ← mod03
  5  Ransomware Exposure ← mod04
  6  Remote Access       ← mod05
  7  User & Process      ← mod06 (hidden if no data)
  8  Cross-Label Matrix  ← mod07
  9  Unmanaged Hosts     ← mod08
  10 Traffic Distribution← mod09
  11 Allowed Traffic     ← mod10
  12 Bandwidth & Volume  ← mod11 (hidden if no data)
  13 Security Findings   ← rules findings
  14 Raw Data            ← df (optional)
"""
from __future__ import annotations

import datetime
import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# ─── Colour palette ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
_SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
_SUBHEADER_FONT = Font(bold=True, size=10)
_ROW_ALT_FILL = PatternFill('solid', fgColor='EBF5FB')
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

_SEVERITY_FILLS = {
    'CRITICAL': PatternFill('solid', fgColor='C0392B'),
    'HIGH':     PatternFill('solid', fgColor='E74C3C'),
    'MEDIUM':   PatternFill('solid', fgColor='F39C12'),
    'LOW':      PatternFill('solid', fgColor='27AE60'),
    'INFO':     PatternFill('solid', fgColor='2980B9'),
}
_SEVERITY_FONTS = {k: Font(color='FFFFFF', bold=True, size=9)
                   for k in _SEVERITY_FILLS}


class ExcelExporter:
    """Export a report results dict to an Excel workbook."""

    def __init__(self, results: dict, df: pd.DataFrame | None = None,
                 include_raw: bool = False):
        self._r = results
        self._df = df
        self._include_raw = include_raw

    def export(self, output_dir: str = 'reports') -> str:
        """Write Excel file and return the full file path."""
        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime('%Y-%m-%d_%H%M')
        filename = f'Illumio_Traffic_Report_{ts}.xlsx'
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        wb.remove(wb.active)   # remove default sheet

        self._sheet_exec_summary(wb)
        self._sheet_mod01(wb)
        self._sheet_mod02(wb)
        self._sheet_mod03(wb)
        self._sheet_mod04(wb)
        self._sheet_mod05(wb)
        self._sheet_mod06(wb)
        self._sheet_mod07(wb)
        self._sheet_mod08(wb)
        self._sheet_mod09(wb)
        self._sheet_mod10(wb)
        self._sheet_mod11(wb)
        self._sheet_mod13(wb)
        self._sheet_mod14(wb)
        self._sheet_mod15(wb)
        self._sheet_findings(wb)
        if self._include_raw and self._df is not None:
            self._sheet_raw(wb)

        wb.save(filepath)
        logger.info(f"[ExcelExporter] Saved: {filepath}")
        return filepath

    # ── helpers ──────────────────────────────────────────────────────────────

    def _new_sheet(self, wb: Workbook, title: str):
        ws = wb.create_sheet(title=title[:31])   # Excel limit 31 chars
        return ws

    def _write_title(self, ws, title: str, row: int = 1, col: int = 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = _TITLE_FONT
        return row + 2

    def _write_df(self, ws, df: pd.DataFrame, start_row: int,
                  start_col: int = 1, color_col: str | None = None) -> int:
        """Write a DataFrame with styled headers. Returns next empty row."""
        if df is None or df.empty:
            ws.cell(row=start_row, column=start_col, value='— No data —')
            return start_row + 2

        # Header row
        for ci, col_name in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=start_row, column=ci, value=str(col_name))
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = _THIN_BORDER

        # Data rows
        for ri, (_, row) in enumerate(df.iterrows(), start=1):
            row_fill = _ROW_ALT_FILL if ri % 2 == 0 else None
            for ci, val in enumerate(row.values, start=start_col):
                cell = ws.cell(row=start_row + ri, column=ci, value=str(val) if not isinstance(val, (int, float)) else val)
                if row_fill:
                    cell.fill = row_fill
                cell.border = _THIN_BORDER
                # Severity colouring
                if color_col and df.columns.tolist()[ci - start_col] == color_col:
                    sev = str(val).upper()
                    if sev in _SEVERITY_FILLS:
                        cell.fill = _SEVERITY_FILLS[sev]
                        cell.font = _SEVERITY_FONTS[sev]

        # Auto-fit column widths
        for ci, col_name in enumerate(df.columns, start=start_col):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 60)

        return start_row + len(df) + 2

    def _write_kv(self, ws, kv_rows: list[dict], start_row: int,
                  key_col: str = 'label', val_col: str = 'value') -> int:
        """Write key-value pairs as a simple 2-column table."""
        for row in kv_rows:
            k_cell = ws.cell(row=start_row, column=1, value=row.get(key_col, ''))
            k_cell.font = Font(bold=True)
            ws.cell(row=start_row, column=2, value=row.get(val_col, ''))
            start_row += 1
        return start_row + 1

    # ── sheets ───────────────────────────────────────────────────────────────

    def _sheet_exec_summary(self, wb: Workbook):
        mod12 = self._r.get('mod12', {})
        ws = self._new_sheet(wb, '📊 Executive Summary')
        row = self._write_title(ws, 'Illumio Traffic Flow Report — Executive Summary')
        ws.cell(row=row, column=1, value=f"Generated: {mod12.get('generated_at', '')}")
        row += 2

        # KPI cards
        ws.cell(row=row, column=1, value='KEY METRICS').font = _SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = _SUBHEADER_FILL
        row += 1
        for kpi in mod12.get('kpis', []):
            row = self._write_kv(ws, [kpi], row)

        row += 1
        # Key findings
        ws.cell(row=row, column=1, value='KEY FINDINGS').font = _SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = _SUBHEADER_FILL
        row += 1
        for kf in mod12.get('key_findings', []):
            ws.cell(row=row, column=1, value=kf.get('severity', '')).fill = \
                _SEVERITY_FILLS.get(kf.get('severity', ''), PatternFill())
            ws.cell(row=row, column=1).font = _SEVERITY_FONTS.get(kf.get('severity', ''), Font())
            ws.cell(row=row, column=2, value=kf.get('finding', ''))
            ws.cell(row=row, column=3, value=kf.get('action', ''))
            row += 1

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60

    def _sheet_mod01(self, wb: Workbook):
        m = self._r.get('mod01', {})
        ws = self._new_sheet(wb, '1 Traffic Overview')
        row = self._write_title(ws, 'Traffic Overview')
        kv = [
            {'label': 'Total Flows', 'value': m.get('total_flows', 0)},
            {'label': 'Total Connections', 'value': m.get('total_connections', 0)},
            {'label': 'Policy Coverage', 'value': f"{m.get('policy_coverage_pct', 0)}%"},
            {'label': 'Allowed', 'value': m.get('allowed_flows', 0)},
            {'label': 'Blocked', 'value': m.get('blocked_flows', 0)},
            {'label': 'Potentially Blocked', 'value': m.get('potentially_blocked_flows', 0)},
            {'label': 'Total Data (MB)', 'value': m.get('total_mb', 0)},
            {'label': 'Date Range', 'value': m.get('date_range', '')},
        ]
        row = self._write_kv(ws, kv, row)
        row = self._write_title(ws, 'Top Ports', row)
        row = self._write_df(ws, m.get('top_ports'), row)
        row = self._write_title(ws, 'Top Protocols', row)
        self._write_df(ws, m.get('top_protocols'), row)

    def _sheet_mod02(self, wb: Workbook):
        m = self._r.get('mod02', {})
        ws = self._new_sheet(wb, '2 Policy Decisions')
        row = self._write_title(ws, 'Policy Decision Breakdown')
        row = self._write_df(ws, m.get('summary'), row)
        for decision in ('allowed', 'blocked', 'potentially_blocked', 'unknown'):
            dm = m.get(decision, {})
            if isinstance(dm, dict) and dm.get('count', 0) > 0:
                label = decision.upper()
                inb = dm.get('inbound_count', 0)
                outb = dm.get('outbound_count', 0)
                row = self._write_kv(ws, [
                    {'label': f'{label} Inbound Flows',  'value': inb},
                    {'label': f'{label} Outbound Flows', 'value': outb},
                ], row)
                row = self._write_title(ws, f"{label} — Top App Flows", row)
                row = self._write_df(ws, dm.get('top_app_flows'), row)
                if inb > 0:
                    row = self._write_title(ws, f"{label} — Top Inbound Ports", row)
                    row = self._write_df(ws, dm.get('top_inbound_ports'), row)
                if outb > 0:
                    row = self._write_title(ws, f"{label} — Top Outbound Ports", row)
                    row = self._write_df(ws, dm.get('top_outbound_ports'), row)
        pc = m.get('port_coverage')
        if pc is not None and hasattr(pc, 'empty') and not pc.empty:
            row = self._write_title(ws, 'Per-Port Coverage', row)
            self._write_df(ws, pc, row)

    def _sheet_mod03(self, wb: Workbook):
        m = self._r.get('mod03', {})
        ws = self._new_sheet(wb, '3 Uncovered Flows')
        row = self._write_title(ws, 'Policy Coverage Gaps')
        kv = [
            {'label': 'Total Uncovered', 'value': m.get('total_uncovered', 0)},
            {'label': 'Policy Coverage %', 'value': f"{m.get('coverage_pct', 0)}%"},
        ]
        inb = m.get('inbound_coverage_pct')
        outb = m.get('outbound_coverage_pct')
        if inb is not None:
            kv.append({'label': 'Inbound Coverage %',  'value': f"{inb}%"})
        if outb is not None:
            kv.append({'label': 'Outbound Coverage %', 'value': f"{outb}%"})
        row = self._write_kv(ws, kv, row)
        row = self._write_title(ws, 'Top Uncovered Flows', row)
        row = self._write_df(ws, m.get('top_flows'), row)
        up = m.get('uncovered_ports')
        if up is not None and hasattr(up, 'empty') and not up.empty:
            row = self._write_title(ws, 'Port Gap Ranking', row)
            row = self._write_df(ws, up, row)
        us = m.get('uncovered_services')
        if us is not None and hasattr(us, 'empty') and not us.empty:
            row = self._write_title(ws, 'Uncovered Services (App + Port)', row)
            row = self._write_df(ws, us, row)
        row = self._write_title(ws, 'By Recommendation Category', row)
        self._write_df(ws, m.get('by_recommendation'), row)

    def _sheet_mod04(self, wb: Workbook):
        m = self._r.get('mod04', {})
        ws = self._new_sheet(wb, '4 Ransomware Exposure')
        row = self._write_title(ws, 'Ransomware Exposure Analysis')
        if 'error' in m:
            ws.cell(row=row, column=1, value=m['error']); return
        ws.cell(row=row, column=1, value=f"Total Risk Flows: {m.get('risk_flows_total', 0)}")
        row += 2
        row = self._write_title(ws, 'Part A: Risk Level Summary', row)
        row = self._write_df(ws, m.get('part_a_summary'), row)
        row = self._write_title(ws, 'Part B: Per-Port Detail', row)
        row = self._write_df(ws, m.get('part_b_per_port'), row, color_col='Risk Level')
        row = self._write_title(ws, 'Part C: By Policy Decision', row)
        row = self._write_df(ws, m.get('part_c_by_decision'), row)
        row = self._write_title(ws, 'Part D: Host Exposure Ranking', row)
        self._write_df(ws, m.get('part_d_host_exposure'), row)

    def _sheet_mod05(self, wb: Workbook):
        m = self._r.get('mod05', {})
        ws = self._new_sheet(wb, '5 Remote Access')
        row = self._write_title(ws, 'Remote Access / Lateral Movement Analysis')
        if m.get('total_lateral_flows', 0) == 0:
            ws.cell(row=row, column=1, value='No lateral movement port traffic found.'); return
        row = self._write_df(ws, m.get('by_service'), row)
        row = self._write_title(ws, 'Top Talkers', row)
        row = self._write_df(ws, m.get('top_talkers'), row)
        row = self._write_title(ws, 'Top Host Pairs', row)
        self._write_df(ws, m.get('top_pairs'), row)

    def _sheet_mod06(self, wb: Workbook):
        m = self._r.get('mod06', {})
        ws = self._new_sheet(wb, '6 User & Process')
        row = self._write_title(ws, 'User & Process Activity')
        if m.get('note'):
            ws.cell(row=row, column=1, value=m['note']); return
        if m.get('user_data_available'):
            row = self._write_df(ws, m.get('top_users'), row)
            row = self._write_title(ws, 'User → Destination App Matrix', row)
            row = self._write_df(ws, m.get('user_dst_matrix'), row)
        if m.get('process_data_available'):
            row = self._write_title(ws, 'Top Processes', row)
            self._write_df(ws, m.get('top_processes'), row)

    def _sheet_mod07(self, wb: Workbook):
        m = self._r.get('mod07', {})
        ws = self._new_sheet(wb, '7 Cross-Label Matrix')
        row = self._write_title(ws, 'Cross-Label Flow Analysis')
        for key, data in m.get('matrices', {}).items():
            row = self._write_title(ws, f'Label Key: {key.upper()}', row)
            if 'note' in data:
                ws.cell(row=row, column=1, value=data['note'])
                row += 2
                continue
            kv = [
                {'label': 'Same-value flows', 'value': data.get('same_value_flows', 0)},
                {'label': 'Cross-value flows', 'value': data.get('cross_value_flows', 0)},
            ]
            row = self._write_kv(ws, kv, row)
            row = self._write_df(ws, data.get('top_cross_pairs'), row)
            row = self._write_title(ws, f'Flow Matrix: {key}', row)
            row = self._write_df(ws, data.get('matrix'), row)

    def _sheet_mod08(self, wb: Workbook):
        m = self._r.get('mod08', {})
        ws = self._new_sheet(wb, '8 Unmanaged Hosts')
        row = self._write_title(ws, 'Unmanaged Host Analysis')
        kv = [
            {'label': 'Unmanaged Flows',        'value': m.get('unmanaged_flow_count', 0)},
            {'label': 'Unmanaged %',             'value': f"{m.get('unmanaged_pct', 0)}%"},
            {'label': 'Unique Unmanaged Src',    'value': m.get('unique_unmanaged_src', 0)},
            {'label': 'Unique Unmanaged Dst',    'value': m.get('unique_unmanaged_dst', 0)},
        ]
        row = self._write_kv(ws, kv, row)
        row = self._write_title(ws, 'Top Unmanaged Sources', row)
        row = self._write_df(ws, m.get('top_unmanaged_src'), row)
        pa = m.get('per_dst_app')
        if pa is not None and hasattr(pa, 'empty') and not pa.empty:
            row = self._write_title(ws, 'Managed Apps Receiving Unmanaged Traffic', row)
            row = self._write_df(ws, pa, row)
        pp = m.get('per_port_proto')
        if pp is not None and hasattr(pp, 'empty') and not pp.empty:
            row = self._write_title(ws, 'Exposed Ports / Protocols', row)
            row = self._write_df(ws, pp, row)
        sp = m.get('src_port_detail')
        if sp is not None and hasattr(sp, 'empty') and not sp.empty:
            row = self._write_title(ws, 'Unmanaged Source × Port Detail', row)
            row = self._write_df(ws, sp, row)
        mh = m.get('managed_hosts_targeted_by_unmanaged')
        if mh is not None and hasattr(mh, 'empty') and not mh.empty:
            row = self._write_title(ws, 'Managed Hosts Targeted by Unmanaged Sources', row)
            self._write_df(ws, mh, row)

    def _sheet_mod09(self, wb: Workbook):
        m = self._r.get('mod09', {})
        ws = self._new_sheet(wb, '9 Traffic Distribution')
        row = self._write_title(ws, 'Traffic Distribution')
        row = self._write_title(ws, 'Port Distribution', row)
        row = self._write_df(ws, m.get('port_distribution'), row)
        row = self._write_title(ws, 'Protocol Distribution', row)
        row = self._write_df(ws, m.get('proto_distribution'), row)
        row = self._write_title(ws, 'Role → Role Flow Matrix', row)
        self._write_df(ws, m.get('role_to_role_matrix'), row)

    def _sheet_mod10(self, wb: Workbook):
        m = self._r.get('mod10', {})
        ws = self._new_sheet(wb, '10 Allowed Traffic')
        row = self._write_title(ws, 'Allowed Traffic Analysis')
        if m.get('note'):
            ws.cell(row=row, column=1, value=m['note']); return
        row = self._write_df(ws, m.get('top_app_flows'), row)
        row = self._write_title(ws, f"Audit Flags (Allowed + Unmanaged Src): {m.get('audit_flag_count', 0)}", row)
        row = self._write_df(ws, m.get('audit_flags'), row)
        row = self._write_title(ws, 'Top Allowed Ports', row)
        self._write_df(ws, m.get('top_allowed_ports'), row)

    def _sheet_mod11(self, wb: Workbook):
        m = self._r.get('mod11', {})
        ws = self._new_sheet(wb, '11 Bandwidth & Volume')
        row = self._write_title(ws, 'Bandwidth & Data Volume Analysis')
        if not m.get('bytes_data_available', False):
            ws.cell(row=row, column=1, value=m.get('note', 'No byte data available.'))
            return
        kv = [
            {'label': 'Total Volume (MB)', 'value': m.get('total_mb', 0)},
            {'label': 'Max Bandwidth (Mbps)', 'value': m.get('max_bandwidth_mbps', 'N/A')},
            {'label': 'P95 Bandwidth (Mbps)', 'value': m.get('p95_bandwidth_mbps', 'N/A')},
        ]
        row = self._write_kv(ws, kv, row)
        row = self._write_title(ws, 'Top Connections by Bytes', row)
        row = self._write_df(ws, m.get('top_by_bytes'), row)
        row = self._write_title(ws, 'Top by Application', row)
        row = self._write_df(ws, m.get('top_app_bytes'), row)
        row = self._write_title(ws, 'Top Bandwidth Connections', row)
        row = self._write_df(ws, m.get('top_bandwidth'), row)
        if 'byte_ratio_anomalies' in m:
            row = self._write_title(ws, 'Byte Ratio Anomalies (potential exfiltration)', row)
            self._write_df(ws, m.get('byte_ratio_anomalies'), row)

    def _sheet_mod13(self, wb: Workbook):
        m = self._r.get('mod13', {})
        ws = self._new_sheet(wb, '13 Enforcement Readiness')
        row = self._write_title(ws, 'Enforcement Readiness Score')
        if 'error' in m:
            ws.cell(row=row, column=1, value=m['error']); return
        row = self._write_kv(ws, [
            {'label': 'Total Score', 'value': f"{m.get('total_score', 0)}/100"},
            {'label': 'Grade',       'value': m.get('grade', '?')},
        ], row)
        ft = m.get('factor_table')
        if ft is not None and hasattr(ft, 'empty') and not ft.empty:
            row = self._write_title(ws, 'Score Breakdown by Factor', row)
            row = self._write_df(ws, ft, row)
        rec = m.get('recommendations')
        if rec is not None and hasattr(rec, 'empty') and not rec.empty:
            row = self._write_title(ws, 'Remediation Recommendations', row)
            self._write_df(ws, rec, row)

    def _sheet_mod14(self, wb: Workbook):
        m = self._r.get('mod14', {})
        ws = self._new_sheet(wb, '14 Infrastructure Scoring')
        row = self._write_title(ws, 'Infrastructure Criticality Analysis')
        if 'error' in m:
            ws.cell(row=row, column=1, value=m['error']); return
        row = self._write_kv(ws, [
            {'label': 'Applications Analysed', 'value': m.get('total_apps', 0)},
            {'label': 'Communication Edges',   'value': m.get('total_edges', 0)},
        ], row)
        rs = m.get('role_summary')
        if rs is not None and hasattr(rs, 'empty') and not rs.empty:
            row = self._write_title(ws, 'Role Distribution', row)
            row = self._write_df(ws, rs, row)
        hub = m.get('hub_apps')
        if hub is not None and hasattr(hub, 'empty') and not hub.empty:
            row = self._write_title(ws, 'Hub Applications (High Blast Radius)', row)
            row = self._write_df(ws, hub, row)
        top = m.get('top_apps')
        if top is not None and hasattr(top, 'empty') and not top.empty:
            row = self._write_title(ws, 'Top Applications by Infrastructure Score', row)
            self._write_df(ws, top, row)

    def _sheet_mod15(self, wb: Workbook):
        m = self._r.get('mod15', {})
        ws = self._new_sheet(wb, '15 Lateral Movement')
        row = self._write_title(ws, 'Lateral Movement Risk Analysis')
        if 'error' in m:
            ws.cell(row=row, column=1, value=m['error']); return
        row = self._write_kv(ws, [
            {'label': 'Lateral Port Flows', 'value': m.get('total_lateral_flows', 0)},
            {'label': 'Lateral %',          'value': f"{m.get('lateral_pct', 0)}%"},
        ], row)
        ss = m.get('service_summary')
        if ss is not None and hasattr(ss, 'empty') and not ss.empty:
            row = self._write_title(ws, 'Lateral Port Activity by Service', row)
            row = self._write_df(ws, ss, row)
        fo = m.get('fan_out_sources')
        if fo is not None and hasattr(fo, 'empty') and not fo.empty:
            row = self._write_title(ws, 'Fan-out Sources (Potential Scanner / Worm)', row)
            row = self._write_df(ws, fo, row)
        al = m.get('allowed_lateral_flows')
        if al is not None and hasattr(al, 'empty') and not al.empty:
            row = self._write_title(ws, 'Explicitly Allowed Lateral Flows', row)
            row = self._write_df(ws, al, row)
        sr = m.get('source_risk_scores')
        if sr is not None and hasattr(sr, 'empty') and not sr.empty:
            row = self._write_title(ws, 'Top High-Risk Sources', row)
            self._write_df(ws, sr, row)

    def _sheet_findings(self, wb: Workbook):
        findings = self._r.get('findings', [])
        ws = self._new_sheet(wb, '🔍 Security Findings')
        row = self._write_title(ws, f'Security Findings ({len(findings)} total)')
        if not findings:
            ws.cell(row=row, column=1, value='No security findings generated.')
            return
        df = pd.DataFrame([f.to_dict() for f in findings])
        self._write_df(ws, df, row, color_col='severity')

    def _sheet_raw(self, wb: Workbook):
        ws = self._new_sheet(wb, 'Raw Data')
        row = self._write_title(ws, f'Raw Data ({len(self._df)} records)')
        # Drop dict columns for Excel compatibility
        safe_df = self._df.copy()
        for col in safe_df.columns:
            if safe_df[col].apply(lambda x: isinstance(x, dict)).any():
                safe_df[col] = safe_df[col].apply(str)
        self._write_df(ws, safe_df, row)
