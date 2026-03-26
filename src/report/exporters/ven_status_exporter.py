"""
src/report/exporters/ven_status_exporter.py
Excel exporter for the VEN Status Inventory Report.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_ORANGE = PatternFill("solid", fgColor="FFEB9C")
_GRAY   = PatternFill("solid", fgColor="D9D9D9")


class VenStatusExporter:
    def __init__(self, results: dict, df: pd.DataFrame = None):
        self.r  = results
        self.df = df

    def export(self, output_dir: str) -> str:
        os.makedirs(output_dir, exist_ok=True)
        filename = f"illumio_ven_status_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        wb.remove(wb.active)

        self._write_summary(wb)
        self._write_sheet(wb, "Online VENs",          self.r.get('online'),         _GREEN,  "375623")
        self._write_sheet(wb, "Offline VENs",         self.r.get('offline'),        _RED,    "9C0006")
        self._write_sheet(wb, "Lost Today (<24h)",    self.r.get('lost_today'),     _RED,    "9C0006")
        self._write_sheet(wb, "Lost Yesterday(24-48h)", self.r.get('lost_yesterday'), _ORANGE, "9C5700")

        wb.save(filepath)
        return filepath

    # ── sheets ────────────────────────────────────────────────────────────────

    def _write_summary(self, wb):
        ws = wb.create_sheet("Executive Summary")
        ws.append(["Illumio VEN Status Inventory Report"])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(["Generated At:", self.r.get('generated_at', '')])
        ws.append([])
        ws.append(["Key Metrics"])
        ws['A4'].font = Font(bold=True)
        for kpi in self.r.get('kpis', []):
            ws.append([kpi['label'], kpi['value']])
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 15

    def _write_sheet(self, wb, title: str, df, fill, font_color: str):
        ws = wb.create_sheet(title)
        count = len(df) if df is not None and not df.empty else 0
        ws.append([f"{title}  ({count})"])
        ws['A1'].font  = Font(bold=True, color=font_color)
        ws['A1'].fill  = fill
        ws.append([])
        self._write_df(ws, df, start_row=3)
        self._autofit(ws)

    # ── helpers ───────────────────────────────────────────────────────────────

    def _write_df(self, ws, df, start_row=1):
        if df is None or df.empty:
            ws.cell(row=start_row, column=1, value="(No records)")
            return
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(
                    lambda x: str(x) if isinstance(x, (dict, list)) else x
                )
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        for cell in ws[start_row]:
            cell.font = Font(bold=True)

    def _autofit(self, ws):
        for col_cells in ws.columns:
            try:
                max_len = max(
                    (len(str(c.value)) for c in col_cells if c.value is not None),
                    default=8,
                )
                ws.column_dimensions[
                    get_column_letter(col_cells[0].column)
                ].width = min(max_len + 2, 45)
            except Exception:
                pass
