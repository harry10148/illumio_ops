"""
src/report/exporters/audit_excel_exporter.py
Excel exporter for Audit Log Reports.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

class AuditExcelExporter:
    def __init__(self, results: dict, df: pd.DataFrame = None):
        self.r = results
        self.df = df

    def export(self, output_dir: str) -> str:
        os.makedirs(output_dir, exist_ok=True)
        filename = f"illumio_audit_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        wb.remove(wb.active)

        self._write_mod00(wb)
        self._write_mod01(wb)
        self._write_mod02(wb)
        self._write_mod03(wb)

        if self.df is not None and not self.df.empty:
            ws = wb.create_sheet("Raw Events")
            self._write_df(ws, self.df.head(10000)) # Limit to 10k for Excel

        wb.save(filepath)
        return filepath

    def _write_mod00(self, wb):
        ws = wb.create_sheet("Executive Summary")
        m = self.r.get('mod00', {})
        ws.append(["Illumio Audit & System Events Report"])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(["Generated At:", m.get('generated_at', '')])
        ws.append([])
        
        ws.append(["Key Metrics"])
        ws['A4'].font = Font(bold=True)
        for kpi in m.get('kpis', []):
            ws.append([kpi['label'], kpi['value']])
        
        ws.append([])
        ws.append(["Top Event Types"])
        top = m.get('top_events_overall', pd.DataFrame())
        if not top.empty:
            self._write_df(ws, top, start_row=ws.max_row + 1)

    def _write_mod01(self, wb):
        ws = wb.create_sheet("System Health & Agent")
        m = self.r.get('mod01', {})
        ws.append(["Total Health Events:", m.get('total_health_events', 0)])
        ws.append([])
        ws.append(["Summary by Event Type"])
        self._write_df(ws, m.get('summary', pd.DataFrame()), start_row=ws.max_row+1)
        ws.append([])
        ws.append(["Recent Health Events"])
        self._write_df(ws, m.get('recent', pd.DataFrame()), start_row=ws.max_row+1)

    def _write_mod02(self, wb):
        ws = wb.create_sheet("User Activity")
        m = self.r.get('mod02', {})
        ws.append(["Total User Events:", m.get('total_user_events', 0)])
        ws.append(["Failed Logins:", m.get('failed_logins', 0)])
        ws.append([])
        ws.append(["Summary by Event Type"])
        self._write_df(ws, m.get('summary', pd.DataFrame()), start_row=ws.max_row+1)
        ws.append([])
        ws.append(["Recent User Events"])
        self._write_df(ws, m.get('recent', pd.DataFrame()), start_row=ws.max_row+1)

    def _write_mod03(self, wb):
        ws = wb.create_sheet("Policy Modifications")
        m = self.r.get('mod03', {})
        ws.append(["Total Policy Events:", m.get('total_policy_events', 0)])
        ws.append([])
        ws.append(["Summary by Event Type"])
        self._write_df(ws, m.get('summary', pd.DataFrame()), start_row=ws.max_row+1)
        ws.append([])
        ws.append(["Recent Policy Events"])
        self._write_df(ws, m.get('recent', pd.DataFrame()), start_row=ws.max_row+1)

    def _write_df(self, ws, df, start_row=1):
        if df is None or df.empty:
            return
        # Stringify any dict/list values so openpyxl can write them
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(
                    lambda x: str(x) if isinstance(x, (dict, list)) else x
                )
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        # Bold header
        for cell in ws[start_row]:
            cell.font = Font(bold=True)
