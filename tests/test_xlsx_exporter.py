"""xlsx exporter tests — openpyxl-based multi-sheet output."""
from __future__ import annotations

import pytest
from openpyxl import load_workbook


@pytest.fixture
def sample_report_result():
    """Minimal ReportResult-shaped dict for xlsx_exporter."""
    return {
        "record_count": 1234,
        "metadata": {
            "title": "Traffic Flow Report",
            "generated_at": "2026-04-18 10:00:00",
            "start_date": "2026-04-11",
            "end_date": "2026-04-18",
        },
        "module_results": {
            "mod01_overview": {
                "summary": "1234 flows analyzed",
                "table": [
                    {"metric": "Total Flows", "value": 1234},
                    {"metric": "Unique Sources", "value": 42},
                ],
            },
            "mod02_policy_decisions": {
                "summary": "",
                "table": [
                    {"decision": "Allowed", "count": 1000},
                    {"decision": "Blocked", "count": 234},
                ],
                "chart_spec": {
                    "type": "pie",
                    "title": "Decisions",
                    "data": {"labels": ["Allowed", "Blocked"], "values": [1000, 234]},
                    "i18n": {"lang": "en"},
                },
            },
        },
    }


def test_xlsx_exporter_creates_workbook(tmp_path, sample_report_result):
    from src.report.exporters.xlsx_exporter import export_xlsx
    out = tmp_path / "report.xlsx"
    export_xlsx(sample_report_result, str(out))
    assert out.exists()
    wb = load_workbook(str(out))
    # Should have one sheet per module plus a summary
    assert "Summary" in wb.sheetnames
    assert any("mod01" in s or "Overview" in s for s in wb.sheetnames)


def test_xlsx_exporter_embeds_chart_image(tmp_path, sample_report_result):
    from src.report.exporters.xlsx_exporter import export_xlsx
    out = tmp_path / "report.xlsx"
    export_xlsx(sample_report_result, str(out))
    wb = load_workbook(str(out))
    # Find the mod02 sheet
    mod02_sheet_name = next((s for s in wb.sheetnames if "mod02" in s or "Policy" in s), None)
    assert mod02_sheet_name is not None
    ws = wb[mod02_sheet_name]
    # openpyxl stores images in ws._images
    assert len(ws._images) >= 1, "chart_spec should produce at least one embedded PNG"


def test_xlsx_exporter_freezes_header_row(tmp_path, sample_report_result):
    from src.report.exporters.xlsx_exporter import export_xlsx
    out = tmp_path / "report.xlsx"
    export_xlsx(sample_report_result, str(out))
    wb = load_workbook(str(out))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fp = ws.freeze_panes
        if fp is None:
            continue  # no data, no freeze needed
        # freeze_panes should be "A{n}" where n >= 2 (header row + 1 or more)
        import re
        assert re.match(r'^A\d+$', fp), f"{sheet_name}: freeze_panes {fp!r} is not an A-column freeze"
        row_num = int(fp[1:])
        assert row_num >= 2, f"{sheet_name}: freeze row {row_num} should be >= 2"


def test_xlsx_exporter_handles_no_chart_spec(tmp_path):
    from src.report.exporters.xlsx_exporter import export_xlsx
    result = {
        "record_count": 10,
        "metadata": {"title": "Minimal"},
        "module_results": {
            "mod_noop": {"summary": "plain", "table": [{"a": 1}]}
        },
    }
    out = tmp_path / "min.xlsx"
    export_xlsx(result, str(out))
    assert out.exists()
