"""Traffic XLSX must contain real DataFrames per sheet, not empty shells."""
import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def sample_flows():
    return pd.DataFrame([
        {"src": "10.0.0.1", "dst": "10.0.0.2", "port": 443, "policy_decision": "allowed"},
        {"src": "10.0.0.1", "dst": "10.0.0.3", "port": 445, "policy_decision": "potentially_blocked"},
        {"src": "10.0.0.4", "dst": "10.0.0.2", "port": 22,  "policy_decision": "blocked"},
    ])


def test_traffic_xlsx_has_expected_sheets(sample_flows, tmp_path):
    from src.report.report_generator import generate_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    generate_traffic_xlsx(sample_flows, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    expected = {"Executive Summary", "Policy Decisions", "Uncovered Flows", "Lateral Movement", "Top Talkers"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"


def test_traffic_xlsx_executive_sheet_has_kpis(sample_flows, tmp_path):
    from src.report.report_generator import generate_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    generate_traffic_xlsx(sample_flows, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    sheet = wb["Executive Summary"]
    rows = list(sheet.iter_rows(values_only=True))
    flat = [str(c) for r in rows for c in r if c is not None]
    assert len(flat) > 0, "Executive Summary sheet is empty"


def test_traffic_xlsx_uncovered_sheet_has_rows(sample_flows, tmp_path):
    from src.report.report_generator import generate_traffic_xlsx
    out_path = tmp_path / "traffic.xlsx"
    generate_traffic_xlsx(sample_flows, str(out_path), profile="security_risk")
    wb = load_workbook(str(out_path))
    assert "Uncovered Flows" in wb.sheetnames
    sheet = wb["Uncovered Flows"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) >= 1, f"Uncovered Flows sheet has no rows"
