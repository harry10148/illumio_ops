"""Audit XLSX must contain real sheets with real rows."""
import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def sample_events():
    return pd.DataFrame([
        {"timestamp": "2026-04-25T01:00:00Z", "actor": "admin", "event": "policy.create",
         "event_type": "policy.create", "status": "success", "resource_type": "rule",
         "created_by": '{"username": "admin"}', "action": '{}', "changes": '{}'},
        {"timestamp": "2026-04-25T02:00:00Z", "actor": "admin", "event": "policy.delete",
         "event_type": "policy.delete", "status": "success", "resource_type": "rule",
         "created_by": '{"username": "admin"}', "action": '{}', "changes": '{}'},
        {"timestamp": "2026-04-25T03:00:00Z", "actor": "svc-deploy", "event": "auth.fail",
         "event_type": "auth.fail", "status": "failed", "resource_type": "",
         "created_by": '{"username": "svc-deploy"}', "action": '{}', "changes": '{}'},
    ])


def test_audit_xlsx_has_expected_sheets(sample_events, tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_events, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Attention Required", "Health", "Users", "Policy Changes", "Correlations"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"


def test_audit_xlsx_has_real_rows(sample_events, tmp_path):
    from src.report.audit_generator import generate_audit_xlsx
    out_path = tmp_path / "audit.xlsx"
    generate_audit_xlsx(sample_events, str(out_path))
    wb = load_workbook(str(out_path))
    sheet = wb["Policy Changes"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) >= 1, "Policy Changes sheet is empty"
