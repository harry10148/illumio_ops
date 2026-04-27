"""VEN XLSX: sheets with real rows."""
from datetime import datetime, timedelta, timezone
import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def sample_workloads():
    now = datetime.now(timezone.utc)
    return pd.DataFrame([
        {"hostname": "web-1", "ven_status": "active",
         "last_heartbeat": now.isoformat()},
        {"hostname": "web-2", "ven_status": "offline",
         "last_heartbeat": (now - timedelta(hours=72)).isoformat()},
        {"hostname": "db-1", "ven_status": "active",
         "last_heartbeat": (now - timedelta(hours=12)).isoformat()},
        {"hostname": "db-2", "ven_status": "active",
         "last_heartbeat": (now - timedelta(hours=36)).isoformat()},
    ])


def test_ven_xlsx_has_expected_sheets(sample_workloads, tmp_path):
    from src.report.ven_status_generator import generate_ven_xlsx
    out_path = tmp_path / "ven.xlsx"
    generate_ven_xlsx(sample_workloads, str(out_path))
    wb = load_workbook(str(out_path))
    expected = {"Online", "Offline", "Lost <24h", "Lost 24-48h"}
    assert expected.issubset(set(wb.sheetnames)), f"missing: {expected - set(wb.sheetnames)}"


def test_ven_xlsx_offline_sheet_has_offline_workloads(sample_workloads, tmp_path):
    from src.report.ven_status_generator import generate_ven_xlsx
    out_path = tmp_path / "ven.xlsx"
    generate_ven_xlsx(sample_workloads, str(out_path))
    wb = load_workbook(str(out_path))
    sheet = wb["Offline"]
    rows = list(sheet.iter_rows(values_only=True))
    flat = [str(c) for r in rows for c in r if c is not None]
    assert any("web-2" in s for s in flat), "web-2 (offline host) not found in Offline sheet"
